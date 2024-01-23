import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl

# 获取网页内容
url = "https://www.shanghairanking.cn/rankings/bcur/202311"
response = requests.get(url)
response.encoding = 'utf-8'

# 使用BeautifulSoup解析网页
soup = BeautifulSoup(response.text, 'html.parser')

# 提取前30的大学排名数据
rankings = []
for row in soup.find_all('tr', limit=31):  # 假设每个排名都在一个 <tr> 标签中
    cols = row.find_all('td')  # <td> 标签中存储了详细信息
    if cols and len(cols) >= 3:  # 检查是否有足够的 <td> 标签
        rank = cols[0].text.strip()
        chinese_name = cols[1].text.strip().split("\n")[0]  # 假设中文名在第一行
        english_name = cols[1].text.strip().split("\n")[1]  # 假设英文名在第二行
        add985_211 = cols[1].text.strip().split("\n")[2]
        address = cols[2].text.strip()
        additional_info = cols[3].text.strip()  # 取得额外信息，如果有的话
        rankings.append({
            'Rank': rank, 
            'Chinese Name': chinese_name, 
            'English Name': english_name, 
            'Add985_211':add985_211,
            'Address':address,
            'Additional Info': additional_info
        })
        if len(rankings) == 30:
            break  # 只提取前30个排名

# 转换为DataFrame
df = pd.DataFrame(rankings)

# 清洗数据：删除前导和尾随的空格
df['English Name'] = df['English Name'].str.strip()
df['Add985_211'] = df['Add985_211'].str.strip()

# 创建一个Pandas Excel writer对象，使用openpyxl作为引擎
with pd.ExcelWriter('formatted_university_rankings.xlsx', engine='openpyxl') as writer:
    # 将DataFrame写入Excel writer对象
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    # 获取openpyxl对象
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # 设置列宽
    worksheet.column_dimensions['A'].width = 5  # 假设排名列宽设为10
    worksheet.column_dimensions['B'].width = 20  # 假设中文名列宽设为20
    worksheet.column_dimensions['C'].width = 58  # 假设英文名列宽设为58
    worksheet.column_dimensions['D'].width = 20  # 假设额外信息列宽设为20
    worksheet.column_dimensions['E'].width = 20  # 假设额外信息列宽设为20
    worksheet.column_dimensions['F'].width = 20  # 假设额外信息列宽设为20

    # 设置行高
    for i in range(1, len(df) + 2):  # +2是因为包括标题行在内，并且从1开始计数
        worksheet.row_dimensions[i].height = 19  # 假设您想要的行高是19

    # 设置Excel单元格的格式
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    #for row in worksheet.iter_rows(min_row=2, max_col=4, max_row=len(df) + 1):
        for cell in row:
            # 设置单元格的对齐方式
            cell.alignment = openpyxl.styles.Alignment(vertical='top', wrap_text=True)
